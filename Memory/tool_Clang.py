from time import sleep

from Utils import  *
import subprocess

class ToolClang:
    def __init__(self, filepath, llvm_path):
        self.file_path = filepath
        self.llvm_path = llvm_path
        self.ccc_analyzer = llvm_path + '/libexec/ccc-analyzer'
        self.clang = llvm_path + '/bin/clang.exe'
        self.compile_output = ''
        self.compile_error = ''
        self.static_scan_output = ''
        self.static_scan_error = ''
        self.run_output = ''
        self.run_error = ''
        self.format_output = ''
        self.format_error = ''
        self.static_scan_strict_output = ''
        self.static_scan_strict_error = ''
        self.static_scan_report_output = ''
        self.static_scan_report_error = ''
        self.code_evaluation_output = ''
        self.code_evaluation_error = ''
        self.is_compile = False

    # clang --driver-mode=gcc -Wall -g3 -o D:/work1/c_test_file/test.exe D:/work1/c_test_file/test.c
    # 编译对应的c文件，并获取对应的错误输出，如果编译成功则输出The program compiles successfully
    def run_compile(self):
        self.init_files(self.file_path.replace('.c', '.exe'))
        # 构建Flawfinder命令
        cmd = ['clang', '--driver-mode=gcc', '-Wall', '-g3', '-o', self.file_path.replace('.c', '.exe'), self.file_path]
        # 执行命令并捕获输出
        try:
            output = subprocess.check_output(cmd, stderr=subprocess.STDOUT, universal_newlines=True)
            print(cmd)
            self.compile_output = output
            if self.compile_output == '':
                self.compile_output = 'The program compiles successfully ( '+self.file_path+' )'

            self.is_compile = True
        except subprocess.CalledProcessError as e:
            self.compile_error = e.output
        print('stdout:')
        print(self.compile_output)
        print('stderr:')
        print(self.compile_error)

    # clang-tidy.exe -checks=*, -header-filter=.*, -extra-arg=-std=c11 D:/work1/c_test_file/test.c
    # ---std=c++98：使用 C++98 标准。
    # -std=c++03：使用 C++03 标准，它是对 C++98 标准的一些修正和更正。
    # -std=c++11：使用 C++11 标准，引入了许多新特性，如自动类型推导、Lambda 表达式、右值引用等。
    # -std=c++14：使用 C++14 标准，是对 C++11 标准的一些改进和扩展。
    # -std=c++17：使用 C++17 标准，引入了许多新特性，如结构化绑定、折叠表达式、std::optional、std::variant 等。
    # -std=c++20：使用 C++20 标准，是对 C++17 标准的一些扩展和改进，包括三路比较运算符、范围 for 循环的初始化语句等。
    # -std=c++23：使用 C++23 标准，这是 C++ 的下一个版本，尚未正式发布，可能会引入新的语言特性和库功能。
    # 对代码进行一个简略的不规范的扫描
    def run_static_scan(self):
        # 构建Flawfinder命令
        cmd = ['clang-tidy', '-checks=*', '-header-filter=.*', '-extra-arg=-std=c11', self.file_path]
        print(cmd)
        # 执行命令并捕获输出
        try:
            output = subprocess.check_output(cmd, stderr=subprocess.STDOUT, universal_newlines=True)
            self.static_scan_output = output

        except subprocess.CalledProcessError as e:
            self.static_scan_error = e.output
            print(self.static_scan_error)
        except Exception as e:
            self.static_scan_error = str(e)
        print('stdout:')
        print(self.static_scan_output)
        print('stderr:')
        print(self.static_scan_error)

    # 运行对应的编译完成的exe文件，需要在函数run_compile之后运行
    def run_exec(self):
        if self.compile_error == '' and os.path.exists(self.file_path.replace('.c', '.exe')):
            try:
                output = subprocess.check_output(self.file_path.replace('.c', '.exe'), shell=True, universal_newlines=True)
                self.run_output = output
            except subprocess.CalledProcessError as e:
                self.run_error = e.stdout
            print('stdout:')
            print(self.run_output)
            print('stderr:')
            print(self.run_error)
    #clang-format.exe -style=LLVM -i D:/work1/c_test_file/test.c
    # LLVM：使用LLVM项目的代码样式规范。
    # Google：使用Google开源项目的代码样式规范。
    # Chromium：使用Chromium项目的代码样式规范。
    # Mozilla：使用Mozilla项目的代码样式规范。
    # WebKit：使用WebKit项目的代码样式规范。
    # GNU：使用GNU项目的代码样式规范。
    # Microsoft：使用Microsoft项目的代码样式规范。
    # 格式化code格式
    def format_code(self):
        cmd = ['clang-format.exe', '-style=LLVM', '-i', self.file_path]
        # 执行命令并捕获输出
        try:
            output = subprocess.check_output(cmd, stderr=subprocess.STDOUT, universal_newlines=True)
            print(cmd)
            self.format_output = output
            if self.format_output == '':
                self.format_output = 'The program reformats successfully ( '+self.file_path+' )'

        except subprocess.CalledProcessError as e:
            self.format_error = e.output
        print('stdout:')
        print(self.format_output)
        print('stderr:')
        print(self.format_error)

    # clang-check.exe -extra-arg=-Wall -extra-arg=-Wextra -extra-arg=-Werror -extra-arg=-pedantic D:/work1/c_test_file/test.c --
    # 执行一个严格的扫描
    def run_static_scan_strict(self):
        # 构建命令
        cmd = ['clang-check.exe', '-extra-arg=-Wall', '-extra-arg=-Wextra', '-extra-arg=-Werror', '-extra-arg=-pedantic', self.file_path]
        print(cmd)
        # 执行命令并捕获输出
        try:
            result = subprocess.run(cmd, capture_output=True, text=True)
            output = result.stderr
            self.static_scan_strict_output = output

        except subprocess.CalledProcessError as e:
            error_output = e.stderr
            self.static_scan_strict_error = error_output
        print('stdout:')
        print(self.static_scan_strict_output)
        print('stderr:')
        print(self.static_scan_strict_error)

    # scan-build.bat --use-cc=X:/llvm/libexec/ccc-analyzer --use-analyzer=X:/llvm/bin/clang.exe clang D:/work1/c_test_file/test.c
    def run_static_scan_report(self):
        # 构建命令
        print(self.ccc_analyzer)
        print(self.clang)
        print(self.file_path)
        cmd = ['scan-build.bat', '--use-cc='+self.ccc_analyzer, '--use-analyzer='+self.clang, 'clang', self.file_path]
        print(cmd)
        # 执行命令并捕获输出
        try:
            result = subprocess.run(cmd, capture_output=True, text=True)
            output = result.stderr
            self.static_scan_report_output = output

        except subprocess.CalledProcessError as e:
            error_output = e.stderr
            self.static_scan_report_error = error_output
        print('stdout:')
        print(self.static_scan_report_output)
        print('stderr:')
        print(self.static_scan_report_error)


    # Filename: 源代码文件的路径。
    # Regions: 区域数，表示代码中被检测到的区域（例如，基本块）的数量。
    # Missed Regions: 未覆盖区域数，表示未被执行的区域数量。
    # Cover: 覆盖率，表示已执行的区域占总区域数的百分比。
    # Functions: 函数数，表示代码中的函数数量。
    # Missed Functions: 未覆盖函数数，表示未被执行的函数数量。
    # Executed Lines: 执行的行数，表示代码中已经执行的行数。
    # Missed Lines: 未执行的行数，表示未被执行的行数。
    # Cover: 行覆盖率，表示已执行的行数占总行数的百分比。
    # Branches: 分支数，表示代码中的分支数量。
    # Missed Branches: 未覆盖分支数，表示未被执行的分支数量。
    # Cover: 分支覆盖率，表示已执行的分支数占总分支数的百分比。
    # cmd1 : clang -fprofile-instr-generate=test.profraw -fcoverage-mapping  D:/work1/c_test_file/test.c -o D:/work1/c_test_file/test.exe
    # cmd2 : llvm-profdata merge -o D:/work1/c_test_file/test.profdata D:/work1/c_test_file/test.profraw
    # cmd3 : llvm-cov report D:\work1\c_test_file\test.exe -instr-profile=D:\work1\c_test_file\test.profdata > D:\work1\c_test_file\test.txt
    def run_code_quality_evaluation(self):
        self.init_files(self.file_path.replace('.c', '.exe'))
        self.init_files(self.file_path.replace('.c', '.profdata'))
        self.init_files(self.file_path.replace('.c', '.profraw'))
        self.init_files(self.file_path.replace('.c', '.xlsx'))
        cmd1 = ['clang', '-fprofile-instr-generate='+self.file_path.replace('.c', '.profraw'), '-fcoverage-mapping', self.file_path, '-o',self.file_path.replace('.c', '.exe')]
        print(cmd1)
        try:
            output = subprocess.check_output(cmd1, stderr=subprocess.STDOUT, universal_newlines=True)
            self.code_evaluation_output = self.code_evaluation_output+output
            # if output == '':
            #     self.code_evaluation_output = self.code_evaluation_output+'The program compiles successfully ( '+self.file_path+' )\n'

        except subprocess.CalledProcessError as e:
            self.code_evaluation_error = self.code_evaluation_error+e.output

        sleep(1)
        self.run_exec()
        cmd2 = ['llvm-profdata', 'merge', '-o', self.file_path.replace('.c', '.profdata'), self.file_path.replace('.c', '.profraw')]
        print(cmd2)
        try:
            output = subprocess.check_output(cmd2, stderr=subprocess.STDOUT, universal_newlines=True)
            self.code_evaluation_output = self.code_evaluation_output + output
            # if output == '':
            #     self.code_evaluation_output = self.code_evaluation_output +'The program generates target parsing file successfully ( '+self.file_path+' )\n'

        except subprocess.CalledProcessError as e:
            self.code_evaluation_error = self.code_evaluation_error+e.output

        sleep(1)
        cmd3 = ['llvm-cov', 'report', self.file_path.replace('.c', '.exe'),
                '-instr-profile='+self.file_path.replace('.c', '.profdata')]
        print(cmd3)
        try:
            # output = subprocess.check_output(cmd3, stderr=subprocess.STDOUT, universal_newlines=True)
            output = subprocess.run(cmd3, capture_output=True, text=True)
            # print('stderr')
            # print(output.stderr)
            # print('stdout')
            # print(output.stdout)
            self.code_evaluation_output = self.code_evaluation_output + output.stdout
        except subprocess.CalledProcessError as e:
            self.code_evaluation_error = self.code_evaluation_error + e.output
        self.from_code_evaluation_get_excel()
        print('stdout:')
        print(self.code_evaluation_output)
        print('stderr:')
        print(self.code_evaluation_error)

    def init_files(self, file_path):
        # 检查文件是否存在
        # print(os.path.exists(file_path))
        if os.path.exists(file_path):
            # 删除文件
            print(file_path+' exists')
            # 检查文件权限
            try:
                # 尝试删除文件
                os.remove(file_path)
                print(f"成功删除文件：{file_path}")
            except PermissionError:
                # 没有写权限，抛出自定义异常
                raise PermissionError(f"没有权限删除文件：{file_path}")
            except FileNotFoundError:
                # 文件不存在，抛出自定义异常
                raise FileNotFoundError(f"文件不存在：{file_path}")
            except Exception as e:
                # 其他异常，打印异常信息
                print(f"发生错误：{str(e)}")

    def from_code_evaluation_get_excel(self):
        import openpyxl

        output_file = self.file_path.replace('.c', '.xlsx')

        # 获取输出内容
        output_content = self.code_evaluation_output

        # 创建Excel工作簿和工作表
        workbook = openpyxl.Workbook()
        worksheet = workbook.active

        # 将输出内容按行分割，并写入工作表中
        lines = output_content.strip().split('\n')
        for row_index, line in enumerate(lines, start=1):
            if row_index == 1:  # 处理表格标题
                columns = [value for value in line.split('  ') if value]  # 剔除空值
                for col_index, value in enumerate(columns, start=1):
                    worksheet.cell(row=row_index, column=col_index).value = value
            else:
                columns = line.split()
                for col_index, value in enumerate(columns, start=1):
                    worksheet.cell(row=row_index, column=col_index).value = value

        # 保存Excel文件
        workbook.save(output_file)
        print(f"Excel文件已生成：{output_file}")

    def from_static_scan_report_get_html(self):
        file_path = get_debug_database() + '\index.html'
        # print(file_path)
        with open(file_path, 'r') as file:
            content = file.read()
        return content


if __name__ == '__main__':
    # #example
    c_file_path = 'D:/work1/c_test_file/test/test.c'
    llvm_path0 = get_available_llvm_path(llvm_path)
    tool_clang = ToolClang(c_file_path, llvm_path0)

    # tool_clang.format_code()

    # tool_clang.run_static_scan_strict()
    # tool_clang.run_compile()
    # tool_clang.run_static_scan()
    # tool_clang.run_exec()
    tool_clang.run_static_scan_report()

    # tool_clang.run_code_quality_evaluation()
    # tool_clang.from_code_evaluation_get_excel()


    #
    # print(tool_clang.compile_output)
    # print(tool_clang.compile_error)
    # print(tool_clang.static_scan_output)
    # print(tool_clang.static_scan_error)
    # ...get other data

    # # 构建命令
    # cmd = ['clang-check', '-extra-arg=-Wall', '-extra-arg=-Wextra', '-extra-arg=-Werror', '-extra-arg=-pedantic',
    #        'D:/work1/c_test_file/test.c']
    #
    # # 执行命令并捕获输出
    # try:
    #     result = subprocess.run(cmd, capture_output=True, text=True)
    #     output = result.stdout
    #     print(output)
    # except subprocess.CalledProcessError as e:
    #     error_output = e.stderr
    #     print(error_output)


